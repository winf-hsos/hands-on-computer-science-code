from openpyxl import load_workbook
from PIL import Image
from tinkerforge.ip_connection import IPConnection
from tinkerforge.bricklet_oled_128x64_v2 import BrickletOLED128x64V2

ipcon = IPConnection()
ipcon.connect("localhost", 4223)
oled = BrickletOLED128x64V2("<YOUR_UID>", ipcon)
oled.clear_display()

workbook = load_workbook("xlsx/Super Mario Pixel Art.xlsx")
sheet = workbook["Super Mario"]

bitmap = []
for row in sheet.iter_rows():
    for cell in row:
        color = getattr(cell.fill.fgColor, "rgb", None)
        color = color[2:] 
        r = int(color[0:2], 16)
        g = int(color[2:4], 16)
        b = int(color[4:6], 16)
        rgb_tupel = (r, g, b)
        bitmap.append(rgb_tupel)

print(f"Bitmap with {len(bitmap)} pixel values")

image = Image.new('RGB', (16, 16))
image.putdata(bitmap)
image.save("bmp/super_mario_color.bmp")