from openpyxl import load_workbook
from PIL import Image
from tinkerforge.ip_connection import IPConnection
from tinkerforge.bricklet_oled_128x64_v2 import BrickletOLED128x64V2

ipcon = IPConnection()
ipcon.connect('localhost', 4223)
oled = BrickletOLED128x64V2('25zo', ipcon)
oled.clear_display()

workbook = load_workbook("xlsx/Darth Vader Pixel Art.xlsx")
sheet = workbook["Darth Vader"]

bits = []
for row in sheet.iter_rows():
    for cell in row:
        color = cell.fill.fgColor.rgb
        if color == "FF000000":
            bits.append(1)
        else:
            bits.append(0)

print(f"Bitmap with {len(bits)} bits: {bits}")

# Saving bitmap to a BMP file
image = Image.new('1', (27, 24))
image.putdata(bits)
image.save("xlsx/darth_vader.bmp")